from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.config import get_settings
from app.models import Record, RecordStageStatus, Role, StageDefinition, User
from app.security import get_password_hash
from app.services.po_status import is_po_received_raw
from app.services.stages import ensure_default_stages, ensure_record_stage_rows, get_active_stages
from app.services.status import derive_status

settings = get_settings()

HEADER_ROW = 2
DATA_START_ROW = 3

ASSIGNEE_EMAIL_DOMAIN = 'cibics.local'

EXCEL_STAGE_MAP = {
    'email sent to customer': 'EMAIL_SENT_TO_CUSTOMER',
    'Data received from Customer': 'DATA_RECEIVED_FROM_CUSTOMER',
    'email sent to bsnl for feasibility': 'EMAIL_SENT_TO_BSNL_FOR_FEASIBILITY',
    'email received from BSNL after feasibility': 'EMAIL_RECEIVED_FROM_BSNL_AFTER_FEASIBILITY',
    'Proposal sent': 'PROPOSAL_SENT',
    'Po received': 'PO_RECEIVED',
}

DUPLICATE_KEY_FIELDS = (
    'sl_no',
    'custodian_code',
    'unlo_code',
    'short_name',
    'custodian_organization',
    'state',
    'site_address',
    'pincode',
)


def _to_text(value: Any) -> str | None:
    if value is None:
        return None
    txt = str(value).strip()
    return txt if txt else None


def _to_bool(value: Any) -> bool:
    if value is None:
        return False
    txt = str(value).strip().lower()
    return txt not in {'', '0', 'false', 'no', 'none'}


def _normalize_name(name: str) -> str:
    return ' '.join(name.replace('\n', ' ').split()).strip()


def _slugify_name(name: str) -> str:
    cleaned = ''.join(ch.lower() if ch.isalnum() else '.' for ch in name)
    cleaned = '.'.join(part for part in cleaned.split('.') if part)
    return cleaned[:40] or 'assignee'


def _is_default_status(value: str | None) -> bool:
    if not value:
        return True
    return is_po_received_raw(value)


def _duplicate_key_from_values(values: dict[str, Any]) -> tuple[str, ...] | None:
    key = tuple((str(values.get(field) or '').strip().lower()) for field in DUPLICATE_KEY_FIELDS)
    return key if any(key) else None


def ensure_assignee_users(db: Session, names: set[str]) -> int:
    created = 0
    existing_names = {
        _normalize_name(name).lower()
        for (name,) in db.query(User.full_name).filter(User.role.in_([Role.ASSIGNEE, Role.SUPER_ADMIN])).all()
    }

    for raw_name in names:
        name = _normalize_name(raw_name)
        if not name or name.lower() in existing_names:
            continue

        base = _slugify_name(name)
        email = f'{base}@{ASSIGNEE_EMAIL_DOMAIN}'
        suffix = 1
        while db.query(User.id).filter(func.lower(User.email) == email.lower()).first():
            suffix += 1
            email = f'{base}{suffix}@{ASSIGNEE_EMAIL_DOMAIN}'

        user = User(
            full_name=name,
            email=email,
            password_hash=get_password_hash(settings.default_assignee_password),
            role=Role.ASSIGNEE,
            is_active=True,
            receive_alert=True,
        )
        db.add(user)
        existing_names.add(name.lower())
        created += 1

    db.flush()
    return created


def _assignee_by_name(db: Session, name: str | None) -> User | None:
    if not name:
        return None
    normalized = _normalize_name(name)
    if not normalized:
        return None

    return (
        db.query(User)
        .filter(User.role.in_([Role.ASSIGNEE, Role.SUPER_ADMIN]))
        .filter(func.lower(User.full_name) == normalized.lower())
        .first()
    )


def _read_rows_from_worksheet(ws: Any) -> list[dict[str, Any]]:
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        value = _to_text(ws.cell(HEADER_ROW, col_idx).value)
        if value:
            headers[col_idx] = value

    expected = {
        'Sl.no',
        'List Type',
        'Type',
        'PO STATUS',
        'Custodian Code',
        'UNLO Code',
        'Short Name',
        'Custodian Organization',
        'State',
        'Site Address',
        'Pincode',
        'Category of Site',
        'Custodian Contact Person Name',
        'Custodian Contact Person Number',
        'Custodian Email',
        'Customer Name',
        'Mobile No',
        'Email id',
    }
    existing_headers = set(headers.values())
    if not expected.issubset(existing_headers):
        missing = sorted(expected - existing_headers)
        raise ValueError(f'Workbook missing expected headers: {missing}')

    header_to_col = {name: idx for idx, name in headers.items()}
    city_col = header_to_col.get('City')
    rows: list[dict[str, Any]] = []

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        row_values = {name: ws.cell(row_idx, idx).value for name, idx in header_to_col.items()}
        if not any(value not in (None, '') for value in row_values.values()):
            continue

        po_status_raw = _to_text(row_values.get('PO STATUS'))
        assignee_hint = _normalize_name(po_status_raw) if po_status_raw and not _is_default_status(po_status_raw) else None
        stage_flags = {
            stage_code: _to_bool(row_values.get(excel_key))
            for excel_key, stage_code in EXCEL_STAGE_MAP.items()
        }
        if is_po_received_raw(po_status_raw):
            stage_flags['PO_RECEIVED'] = True

        rows.append(
            {
                'source_row': row_idx,
                'sl_no': _to_text(row_values.get('Sl.no')),
                'list_type': _to_text(row_values.get('List Type')),
                'type': _to_text(row_values.get('Type')),
                'po_status_raw': po_status_raw,
                'custodian_code': _to_text(row_values.get('Custodian Code')),
                'unlo_code': _to_text(row_values.get('UNLO Code')),
                'short_name': _to_text(row_values.get('Short Name')),
                'custodian_organization': _to_text(row_values.get('Custodian Organization')),
                'state': _to_text(row_values.get('State')),
                'site_address': _to_text(row_values.get('Site Address')),
                'city': _to_text(ws.cell(row_idx, city_col).value) if city_col else _to_text(ws.cell(row_idx, 12).value),
                'pincode': _to_text(row_values.get('Pincode')),
                'category_of_site': _to_text(row_values.get('Category of Site')),
                'custodian_contact_person_name': _to_text(row_values.get('Custodian Contact Person Name')),
                'custodian_contact_person_number': _to_text(row_values.get('Custodian Contact Person Number')),
                'custodian_email': _to_text(row_values.get('Custodian Email')),
                'customer_name': _to_text(row_values.get('Customer Name')),
                'mobile_no': _to_text(row_values.get('Mobile No')),
                'client_email': _to_text(row_values.get('Email id')),
                'assignee_name_hint': assignee_hint,
                'stage_flags': stage_flags,
            }
        )

    return rows


def _analyze_rows(db: Session, rows: list[dict[str, Any]]) -> dict[str, Any]:
    existing_rows = db.query(Record).all()
    existing_source_rows = {item.source_row for item in existing_rows}
    existing_keys: set[tuple[str, ...]] = set()
    for item in existing_rows:
        key = _duplicate_key_from_values(
            {
                'sl_no': item.sl_no,
                'custodian_code': item.custodian_code,
                'unlo_code': item.unlo_code,
                'short_name': item.short_name,
                'custodian_organization': item.custodian_organization,
                'state': item.state,
                'site_address': item.site_address,
                'pincode': item.pincode,
            }
        )
        if key:
            existing_keys.add(key)

    seen_upload_keys: set[tuple[str, ...]] = set()
    duplicate_rows = 0
    analyzed_rows: list[dict[str, Any]] = []
    for row in rows:
        reasons: list[str] = []
        if row['source_row'] in existing_source_rows:
            reasons.append('EXISTING_SOURCE_ROW')

        dedupe_key = _duplicate_key_from_values(row)
        if dedupe_key and dedupe_key in existing_keys:
            reasons.append('EXISTING_DATA')
        if dedupe_key and dedupe_key in seen_upload_keys:
            reasons.append('DUPLICATE_IN_FILE')
        if dedupe_key:
            seen_upload_keys.add(dedupe_key)

        duplicate = len(reasons) > 0
        if duplicate:
            duplicate_rows += 1
        analyzed_rows.append({**row, 'duplicate': duplicate, 'duplicate_reasons': reasons})

    return {
        'rows': analyzed_rows,
        'total_rows': len(rows),
        'duplicate_rows': duplicate_rows,
        'insertable_rows': len(rows) - duplicate_rows,
    }


def analyze_phoneme_excel_bytes(db: Session, file_bytes: bytes, preview_limit: int = 200) -> dict[str, Any]:
    ensure_default_stages(db)
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active
    analyzed = _analyze_rows(db, _read_rows_from_worksheet(ws))

    preview_rows = [
        {
            'source_row': row['source_row'],
            'sl_no': row['sl_no'],
            'short_name': row['short_name'],
            'custodian_organization': row['custodian_organization'],
            'state': row['state'],
            'custodian_code': row['custodian_code'],
            'unlo_code': row['unlo_code'],
            'duplicate': row['duplicate'],
            'duplicate_reasons': row['duplicate_reasons'],
        }
        for row in analyzed['rows'][: max(1, preview_limit)]
    ]
    return {
        'total_rows': analyzed['total_rows'],
        'duplicate_rows': analyzed['duplicate_rows'],
        'insertable_rows': analyzed['insertable_rows'],
        'preview_rows': preview_rows,
    }


def import_phoneme_excel_bytes(db: Session, file_bytes: bytes) -> dict[str, int]:
    ensure_default_stages(db)
    active_stages = get_active_stages(db)
    stage_by_code = {item.code: item for item in active_stages}

    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active
    analyzed = _analyze_rows(db, _read_rows_from_worksheet(ws))

    assignee_names = set()
    created_records: list[Record] = []
    created = 0
    for row in analyzed['rows']:
        if row['duplicate']:
            continue

        assignee_hint = row.get('assignee_name_hint')
        if assignee_hint:
            assignee_names.add(assignee_hint)

        record = Record(
            source_row=row['source_row'],
            sl_no=row['sl_no'],
            list_type=row['list_type'],
            type=row['type'],
            po_status_raw=row['po_status_raw'],
            custodian_code=row['custodian_code'],
            unlo_code=row['unlo_code'],
            short_name=row['short_name'],
            custodian_organization=row['custodian_organization'],
            state=row['state'],
            site_address=row['site_address'],
            city=row['city'],
            pincode=row['pincode'],
            category_of_site=row['category_of_site'],
            custodian_contact_person_name=row['custodian_contact_person_name'],
            custodian_contact_person_number=row['custodian_contact_person_number'],
            custodian_email=row['custodian_email'],
            customer_name=row['customer_name'],
            mobile_no=row['mobile_no'],
            client_email=row['client_email'],
            assignee_name_hint=assignee_hint,
            assignee_id=None,
            email_alert_pending=False,
            last_email_alert_at=None,
        )
        db.add(record)
        db.flush()
        created += 1
        created_records.append(record)

        ensure_record_stage_rows(db, record, active_stages)
        stage_rows = {
            item.stage.code: item
            for item in (
                db.query(RecordStageStatus)
                .join(StageDefinition, StageDefinition.id == RecordStageStatus.stage_id)
                .filter(RecordStageStatus.record_id == record.id, StageDefinition.is_active.is_(True))
                .all()
            )
        }

        for stage_code, completed in row['stage_flags'].items():
            stage_obj = stage_by_code.get(stage_code)
            stage_row = stage_rows.get(stage_code)
            if not stage_obj or not stage_row:
                continue
            stage_row.is_completed = completed
            stage_row.completed_at = datetime.now(timezone.utc) if completed else None
            stage_row.notes = None

        record.status = derive_status(record)

    assignees_created = ensure_assignee_users(db, assignee_names)
    for item in created_records:
        if item.assignee_name_hint:
            assignee = _assignee_by_name(db, item.assignee_name_hint)
            item.assignee_id = assignee.id if assignee else None

    db.commit()

    return {
        'imported': created,
        'created': created,
        'updated': 0,
        'assignees_created': assignees_created,
        'skipped_duplicates': analyzed['duplicate_rows'],
    }


def import_phoneme_excel(db: Session, filepath: str) -> dict[str, int]:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f'File not found: {path}')
    with path.open('rb') as fh:
        return import_phoneme_excel_bytes(db, fh.read())
